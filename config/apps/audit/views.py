import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from rest_framework import filters, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response
from rest_framework.viewsets import ReadOnlyModelViewSet

from .models import AuditLog
from .serializers import AuditLogSerializer, AuditLogListSerializer


class AuditLogViewSet(ReadOnlyModelViewSet):
    """
    Read-only viewset — only admin users can access audit logs.

    Supports filtering via query params:
        ?module=Inventory
        ?action=UPDATE
        ?user=john
        ?date_from=2026-01-01
        ?date_to=2026-03-31
        ?search=<term>        — searches username, record_repr, record_id
        ?page=1&page_size=50
    """

    queryset           = AuditLog.objects.select_related("user").all()
    permission_classes = [IsAdminUser]
    filter_backends    = [filters.SearchFilter, filters.OrderingFilter]
    search_fields      = ["username_snapshot", "full_name_snapshot",
                          "record_repr", "record_id", "ip_address"]
    ordering_fields    = ["timestamp", "module", "action", "username_snapshot"]
    ordering           = ["-timestamp"]

    def get_serializer_class(self):
        if self.action == "list":
            return AuditLogListSerializer
        return AuditLogSerializer

    # ── Filtering ─────────────────────────────────────────────────────────────

    def get_queryset(self):
        qs     = super().get_queryset()
        params = self.request.query_params

        module    = params.get("module")
        action_p  = params.get("action")
        user_q    = params.get("user", "").strip()
        date_from = params.get("date_from")
        date_to   = params.get("date_to")

        if module:
            qs = qs.filter(module=module)
        if action_p:
            qs = qs.filter(action=action_p.upper())
        if user_q:
            qs = qs.filter(username_snapshot__icontains=user_q)
        if date_from:
            d = parse_date(date_from)
            if d:
                qs = qs.filter(timestamp__date__gte=d)
        if date_to:
            d = parse_date(date_to)
            if d:
                qs = qs.filter(timestamp__date__lte=d)

        return qs

    # ── Summary stats ─────────────────────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="summary")
    def summary(self, request):
        """
        Returns aggregate counts for the dashboard cards.
        Optionally filtered by the same query params as the list view.
        """
        from django.db.models import Count

        qs = self.get_queryset()

        by_action = (
            qs.values("action")
              .annotate(count=Count("id"))
              .order_by("-count")
        )
        by_module = (
            qs.values("module")
              .annotate(count=Count("id"))
              .order_by("-count")
        )
        by_user = (
            qs.values("username_snapshot")
              .annotate(count=Count("id"))
              .order_by("-count")[:10]
        )

        return Response({
            "total":     qs.count(),
            "by_action": list(by_action),
            "by_module": list(by_module),
            "by_user":   list(by_user),
        })

    # ── Excel export ──────────────────────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        """
        Export filtered logs to Excel (.xlsx).
        Respects the same query-param filters as the list view.
        Capped at 10 000 rows to protect server memory.
        """
        qs = self.filter_queryset(self.get_queryset())[:10_000]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit Logs"

        # ── Header row ────────────────────────────────────────────────────────
        headers = [
            "ID", "Timestamp", "User ID", "Username", "Full Name",
            "Module", "Action", "Record ID", "Record",
            "Changed Fields", "IP Address", "User Agent",
        ]
        header_fill = PatternFill(
            start_color="1A6EF5", end_color="1A6EF5", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 22

        # ── Data rows ─────────────────────────────────────────────────────────
        for row_num, log in enumerate(qs, start=2):
            changed = (
                ", ".join(log.changed_fields)
                if log.changed_fields
                else ""
            )
            ts = log.timestamp
            if ts and timezone.is_aware(ts):
                ts = ts.astimezone(timezone.get_current_timezone())

            ws.append([
                log.id,
                ts.strftime("%Y-%m-%d %H:%M:%S") if ts else "",
                log.user_id_snapshot or "",
                log.username_snapshot,
                log.full_name_snapshot,
                log.module,
                log.action,
                log.record_id,
                log.record_repr,
                changed,
                log.ip_address or "",
                log.user_agent[:200] if log.user_agent else "",
            ])

            # Alternate row shading
            if row_num % 2 == 0:
                fill = PatternFill(
                    start_color="F0F4FF", end_color="F0F4FF", fill_type="solid"
                )
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).fill = fill

        # ── Column widths ─────────────────────────────────────────────────────
        col_widths = [8, 20, 9, 18, 22, 14, 10, 12, 35, 28, 16, 30]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = width

        # ── Stream response ───────────────────────────────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"audit_logs_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    # ── Choices (for filter dropdowns) ────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="choices")
    def choices(self, request):
        return Response({
            "modules": [
                {"value": v, "label": l}
                for v, l in AuditLog.MODULE_CHOICES
            ],
            "actions": [
                {"value": v, "label": l}
                for v, l in AuditLog.ACTION_CHOICES
            ],
        })
